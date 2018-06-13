import openpyxl
wb = openpyxl.load_workbook('EH_Payroll_Needed.xlsx')
ws = wb.get_sheet_by_name('Export Worksheet')
 
rows = []
datafiles = []
adict = {}
for row in range(2, ws.max_row + 1):
    file_name = ws['A' + str(row)].value
    rownum = ws['B' + str(row)].value
    try:
        adict[file_name].append(rownum)
    except:
        adict[file_name] = [rownum]
#print(adict)

for k,v in adict.items():
    newdict = {}
    newdict['file'] = k
    newdict['rows'] = v 
    datafiles.append(newdict)

print(datafiles)     
        
        
        
        
        
        
        
        
        
